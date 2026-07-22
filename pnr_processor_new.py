import re
import requests
from xml.etree import ElementTree as ET
from typing import Optional
import argparse
import openpyxl
from openpyxl import load_workbook
import time

# 全局PNR变量
current_pnr = "HZMZ28"  # 默认值


def set_pnr(pnr: str):
    global current_pnr
    current_pnr = pnr


def get_ckin_info(booking_reference_id: Optional[str] = None, max_retries: int = 3) -> Optional[str]:
    """获取CKIN信息，如果不传booking_reference_id则使用current_pnr"""
    if booking_reference_id is None:
        booking_reference_id = current_pnr
    url = 'https://eterm1w-api.etriplink.com/ibe-proxy/ota/xml/AirResRetRtl'
    headers = {
        'Aid': '84403',
        'Content-Type': 'application/xml',
        'Authorization': 'Basic a21nMzE5enN5OnludGIxMjU4MA==',
        'Cookie': 'X-LB=2.3c5.e609bae1.50'
    }

    xml_data = f'''<?xml version="1.0" encoding="UTF-8"?>
<OTA_AirResRetRtlRQ xmlns="http://espeed.travelsky.com"
xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" EchoToken="String"
TimeStamp="String" Version="String" Target="String">
<POS>
<Source PseudoCityCode="KMG319" />
</POS>
<BookingReferenceID ID="{booking_reference_id}">
</BookingReferenceID>
</OTA_AirResRetRtlRQ>'''

    for attempt in range(max_retries):
        try:
            response = requests.post(url, headers=headers, data=xml_data, timeout=10)
            response.raise_for_status()
            root = ET.fromstring(response.text)

            # 查找所有包含CKIN的Text节点
            for special_service in root.findall('.//SpecialServiceRequest'):
                if special_service.get('SSRCode') == 'CKIN':
                    text_node = special_service.find('Text')
                    if text_node is not None and text_node.text is not None and re.search(r'VICO\d{2}GP\d{3}', text_node.text):
                        return text_node.text
            return None

        except (requests.RequestException, ET.ParseError) as e:
            if attempt == max_retries - 1:
                print(f"查询PNR {booking_reference_id} 失败: {str(e)}")
                return None
            time.sleep(1)

    return None


def get_rtu1_info(booking_reference_id: Optional[str] = None, extract_variable: str = None, max_retries: int = 3) -> dict:
    """
    使用RTU1指令查询PNR的EI/EI*GP项目信息
    RTU1用于查询PNR中的航段、票价等信息

    参数:
        booking_reference_id: PNR号
        extract_variable: 如果提供，会在响应中搜索并提取该变量
    """
    if booking_reference_id is None:
        booking_reference_id = current_pnr

    result = {
        "ei_info": None,
        "ei_gp_info": None,
        "raw_response": None,
        "extracted_variable": None
    }

    # 先获取完整的PNR信息
    url = 'https://eterm1w-api.etriplink.com/ibe-proxy/ota/xml/AirResRetRtl'
    headers = {
        'Aid': '84403',
        'Content-Type': 'application/xml',
        'Authorization': 'Basic a21nMzE5enN5OnludGIxMjU4MA==',
        'Cookie': 'X-LB=2.3c5.e609bae1.50'
    }

    xml_data = f'''<?xml version="1.0" encoding="UTF-8"?>
<OTA_AirResRetRtlRQ xmlns="http://espeed.travelsky.com"
xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" EchoToken="String"
TimeStamp="String" Version="String" Target="String">
<POS>
<Source PseudoCityCode="KMG319" />
</POS>
<BookingReferenceID ID="{booking_reference_id}">
</BookingReferenceID>
</OTA_AirResRetRtlRQ>'''

    for attempt in range(max_retries):
        try:
            response = requests.post(url, headers=headers, data=xml_data, timeout=10)
            response.raise_for_status()
            response_text = response.text
            result["raw_response"] = response_text

            print(f"\n步骤1: 查询PNR {booking_reference_id} 成功")
            print(f"=" * 80)

            # 解析EI和EI*GP信息
            root = ET.fromstring(response_text)

            # 查找所有SpecialRemark节点
            for special_remark in root.findall('.//SpecialRemark'):
                remark_text = special_remark.find('Text')
                if remark_text is not None and remark_text.text:
                    text = remark_text.text
                    # 匹配包含GP的FARECODE (EI*GP)
                    if 'FARECODE/GP' in text:
                        result["ei_gp_info"] = text
                        print(f"\n[OK] 找到EI*GP信息(FARECODE): {result['ei_gp_info']}")
                    # 匹配包含EI的备注 (标准EI)
                    elif text.startswith('EI') or re.match(r'^EI\s', text):
                        if not result["ei_info"]:  # 只保存第一个EI信息
                            result["ei_info"] = text
                            print(f"\n[OK] 找到EI信息: {result['ei_info']}")

            # 如果需要提取特定变量
                if extract_variable:
                    print(f"\n步骤2: 提取变量 '{extract_variable}'")
                    print(f"=" * 80)

                    # 在整个响应中搜索变量
                    if extract_variable in response_text:
                        print(f"\n[OK] 找到变量 '{extract_variable}'")

                    # 尝试提取完整的上下文
                    all_text_content = []

                    # 从SpecialRemark中提取
                    for special_remark in root.findall('.//SpecialRemark'):
                        remark_text = special_remark.find('Text')
                        if remark_text is not None and remark_text.text and extract_variable in remark_text.text:
                            all_text_content.append({
                                'location': 'SpecialRemark',
                                'content': remark_text.text
                            })
                            print(f"\n  位置: SpecialRemark")
                            print(f"  内容: {remark_text.text}")

                    # 从OtherServiceInformation中提取
                    for osi in root.findall('.//OtherServiceInformation'):
                        osi_text = osi.find('Text')
                        if osi_text is not None and osi_text.text and extract_variable in osi_text.text:
                            all_text_content.append({
                                'location': 'OtherServiceInformation',
                                'content': osi_text.text
                            })
                            print(f"\n  位置: OtherServiceInformation")
                            print(f"  内容: {osi_text.text}")

                    # 从SpecialServiceRequest中提取
                    for ssr in root.findall('.//SpecialServiceRequest'):
                        ssr_text = ssr.find('Text')
                        if ssr_text is not None and ssr_text.text and extract_variable in ssr_text.text:
                            all_text_content.append({
                                'location': 'SpecialServiceRequest',
                                'content': ssr_text.text
                            })
                            print(f"\n  位置: SpecialServiceRequest")
                            print(f"  内容: {ssr_text.text}")

                    if all_text_content:
                        result["extracted_variable"] = {
                            'variable': extract_variable,
                            'found': True,
                            'locations': all_text_content
                        }
                        print(f"\n[OK] 成功提取变量 '{extract_variable}'，共找到 {len(all_text_content)} 处")
                    else:
                        print(f"\n[WARN] 变量 '{extract_variable}' 存在于响应中，但不在标准字段中")
                        result["extracted_variable"] = {
                            'variable': extract_variable,
                            'found': True,
                            'in_raw_response': True
                        }
                else:
                    print(f"\n[FAIL] 未找到变量 '{extract_variable}'")
                    result["extracted_variable"] = {
                        'variable': extract_variable,
                        'found': False
                    }

            # 总结查询结果
            print(f"\n步骤3: 查询结果总结")
            print(f"=" * 80)
            if result["ei_info"]:
                print(f"EI信息: {result['ei_info']}")
            else:
                print(f"EI信息: 未找到")
            if result["ei_gp_info"]:
                print(f"EI*GP信息: {result['ei_gp_info']}")
            else:
                print(f"EI*GP信息: 未找到")
            print(f"=" * 80)

            # 如果成功解析，返回结果
            if result["ei_info"] or result["ei_gp_info"] or (extract_variable and result["extracted_variable"] and result["extracted_variable"]["found"]):
                return result

        except (requests.RequestException, ET.ParseError) as e:
            print(f"查询失败: {str(e)}")
            if attempt == max_retries - 1:
                continue
            time.sleep(1)

    print(f"\n无法获取 {booking_reference_id} 的EI/EI*GP信息")
    return result


def get_tn_info(booking_reference_id: Optional[str] = None, ticket_number: Optional[str] = None, max_retries: int = 3) -> dict:
    """
    使用DETR TN指令查询旅客姓名和出票时间
    如果提供ticket_number则使用票号查询，否则使用PNR查询
    """
    # 尝试多个可能的API端点
    endpoints = [
        {
            'url': 'https://eterm1w-api.etriplink.com/ibe-proxy/command',
            'type': 'json',
            'headers': {
                'Aid': '84403',
                'Content-Type': 'application/json',
                'Authorization': 'Basic a21nMzE5enN5OnludGIxMjU4MA==',
                'Cookie': 'X-LB=2.3c5.e609bae1.50'
            }
        },
        {
            'url': 'https://eterm1w-api.etriplink.com/ibe-proxy/ota/xml/AirResRetRtl',
            'type': 'xml',
            'headers': {
                'Aid': '84403',
                'Content-Type': 'application/xml',
                'Authorization': 'Basic a21nMzE5enN5OnludGIxMjU4MA==',
                'Cookie': 'X-LB=2.3c5.e609bae1.50'
            }
        }
    ]

    # 根据参数决定查询方式
    if ticket_number:
        command = f"DETR TN/{ticket_number}"
        query_id = ticket_number
    elif booking_reference_id:
        command = f"DETR TN/{booking_reference_id}"
        query_id = booking_reference_id
    else:
        command = f"DETR TN/{current_pnr}"
        query_id = current_pnr

    result = {
        "passenger_name": None,
        "ticket_time": None,
        "raw_response": None
    }

    for endpoint in endpoints:
        url = endpoint['url']
        headers = endpoint['headers']
        endpoint_type = endpoint['type']

        for attempt in range(max_retries):
            try:
                print(f"\n尝试使用URL: {url} (类型: {endpoint_type})")

                if endpoint_type == 'json':
                    data = {
                        "command": command,
                        "pseudoCityCode": "KMG319"
                    }
                    response = requests.post(url, headers=headers, json=data, timeout=10)
                elif endpoint_type == 'xml':
                    # 尝试使用OTACmd格式的XML
                    xml_data = f'''<?xml version="1.0" encoding="UTF-8"?>
<OTACmdRQ xmlns="http://espeed.travelsky.com">
<POS><Source PseudoCityCode="KMG319"/></POS>
<Command>{command}</Command>
</OTACmdRQ>'''
                    response = requests.post(url, headers=headers, data=xml_data, timeout=10)

                # 如果返回HTML错误页面，尝试下一个端点
                if '<html>' in response.text.lower():
                    print(f"返回HTML错误，尝试下一个端点...")
                    break

                response.raise_for_status()
                response_text = response.text
                result["raw_response"] = response_text

                print(f"\n原始响应内容:")
                print(response_text)

                # 解析旅客姓名
                name_match = re.search(r'([A-Z]{2}/[A-Z]{2,20})', response_text)
                if name_match:
                    result["passenger_name"] = name_match.group(1)

                # 解析出票时间，格式如：05JAN26/14:30
                time_match = re.search(r'(\d{2}[A-Z]{3}\d{2}/\d{2}:\d{2})', response_text)
                if time_match:
                    result["ticket_time"] = time_match.group(1)

                # 如果没有找到出票时间，尝试其他格式
                if not result["ticket_time"]:
                    alt_time_match = re.search(r'(\d{2}\D{3}\d{2}\s+\d{2}:\d{2})', response_text)
                    if alt_time_match:
                        result["ticket_time"] = alt_time_match.group(1)

                # 如果成功解析或得到有效响应，直接返回
                if result["passenger_name"] or result["ticket_time"]:
                    print(f"\n[OK] 成功获取到信息！")
                    return result

            except (requests.RequestException, Exception) as e:
                print(f"查询失败: {str(e)}")
                if attempt == max_retries - 1:
                    continue
                time.sleep(1)

    print(f"\n所有API端点均无法获取 {query_id} 的TN信息")
    print("可能的原因:")
    print("1. API端点不存在或不支持DETR指令")
    print("2. 需要使用不同的认证方式")
    print("3. 需要使用其他接口来查询票号信息")
    return result



def process_excel(input_file: str, output_file: str, query_type: str = "ckin"):
    """
    处理Excel文件，查询PNR并写入结果
    query_type: 'ckin' 查询CKIN信息, 'tn' 查询TN信息（旅客姓名和出票时间）, 'rtu1' 查询RTU1信息（EI/EI*GP）, 'all' 查询所有信息
    """
    try:
        wb = load_workbook(input_file)
        ws = wb.active
        total_rows = ws.max_row - 1
        processed = 0

        for row in ws.iter_rows(min_row=2):  # 从第二行开始
            pnr = row[0].value
            if pnr:
                if query_type == "ckin":
                    ckin_info = get_ckin_info(pnr)
                    row[1].value = ckin_info if ckin_info else "未找到CKIN信息"
                elif query_type == "tn":
                    tn_info = get_tn_info(pnr)
                    row[1].value = tn_info["passenger_name"] if tn_info["passenger_name"] else "未找到旅客姓名"
                    row[2].value = tn_info["ticket_time"] if tn_info["ticket_time"] else "未找到出票时间"
                elif query_type == "rtu1":
                    rtu1_info = get_rtu1_info(pnr)
                    row[1].value = rtu1_info["ei_info"] if rtu1_info["ei_info"] else "未找到EI信息"
                    row[2].value = rtu1_info["ei_gp_info"] if rtu1_info["ei_gp_info"] else "未找到EI*GP信息"
                elif query_type == "all":
                    ckin_info = get_ckin_info(pnr)
                    tn_info = get_tn_info(pnr)
                    rtu1_info = get_rtu1_info(pnr)
                    row[1].value = ckin_info if ckin_info else "未找到CKIN信息"
                    row[2].value = tn_info["passenger_name"] if tn_info["passenger_name"] else "未找到旅客姓名"
                    row[3].value = tn_info["ticket_time"] if tn_info["ticket_time"] else "未找到出票时间"
                    row[4].value = rtu1_info["ei_info"] if rtu1_info["ei_info"] else "未找到EI信息"
                    row[5].value = rtu1_info["ei_gp_info"] if rtu1_info["ei_gp_info"] else "未找到EI*GP信息"

                processed += 1
                print(f"\r处理进度: {processed}/{total_rows} ({processed/total_rows:.1%})", end="")

        wb.save(output_file)
        print(f"\n处理完成，结果已保存到 {output_file}")
        return True
    except Exception as e:
        print(f"\n处理Excel时出错: {str(e)}")
        return False


def main():
    parser = argparse.ArgumentParser(description='获取CKIN、TN和RTU1信息')
    parser.add_argument('--pnr', help='设置PNR号')
    parser.add_argument('--ticket', help='票号（用于TN查询）')
    parser.add_argument('--get-ckin', action='store_true', help='获取当前CKIN信息')
    parser.add_argument('--get-tn', action='store_true', help='获取当前TN信息（旅客姓名和出票时间）')
    parser.add_argument('--get-rtu1', action='store_true', help='获取当前RTU1信息（EI/EI*GP）')
    parser.add_argument('--extract', help='从PNR中提取指定的变量（配合--get-rtu1使用）')
    parser.add_argument('--input', help='输入Excel文件路径')
    parser.add_argument('--output', help='输出Excel文件路径')
    parser.add_argument('--type', choices=['ckin', 'tn', 'rtu1', 'all'], default='ckin',
                        help='查询类型: ckin(CKIN信息), tn(旅客姓名和出票时间), rtu1(EI/EI*GP信息), all(全部信息)')

    args = parser.parse_args()

    if args.pnr:
        set_pnr(args.pnr)

    if args.get_ckin:
        ckin_info = get_ckin_info()
        print(f"当前CKIN信息: {ckin_info}" if ckin_info else "未找到CKIN信息")
    elif args.get_tn:
        # 支持使用票号或PNR查询
        tn_info = get_tn_info(ticket_number=args.ticket)
        print(f"旅客姓名: {tn_info['passenger_name']}" if tn_info['passenger_name'] else "未找到旅客姓名")
        print(f"出票时间: {tn_info['ticket_time']}" if tn_info['ticket_time'] else "未找到出票时间")
    elif args.get_rtu1:
        # 获取RTU1信息（EI/EI*GP）
        rtu1_info = get_rtu1_info(extract_variable=args.extract)
        print(f"EI信息: {rtu1_info['ei_info']}" if rtu1_info['ei_info'] else "未找到EI信息")
        print(f"EI*GP信息: {rtu1_info['ei_gp_info']}" if rtu1_info['ei_gp_info'] else "未找到EI*GP信息")
        if args.extract and rtu1_info['extracted_variable']:
            var_info = rtu1_info['extracted_variable']
            if var_info['found']:
                print(f"\n提取变量 '{args.extract}': 找到")
            else:
                print(f"\n提取变量 '{args.extract}': 未找到")
    elif args.input and args.output:
        process_excel(args.input, args.output, args.type)
    else:
        print("使用说明:")
        print("  --pnr PNR号: 设置PNR号")
        print("  --ticket 票号: 设置票号（用于TN查询）")
        print("  --get-ckin: 获取CKIN信息")
        print("  --get-tn: 获取TN信息（旅客姓名和出票时间）")
        print("  --get-rtu1: 获取RTU1信息（EI/EI*GP）")
        print("  --extract 变量名: 从PNR中提取指定变量（配合--get-rtu1使用）")
        print("  --input 输入文件 --output 输出文件: 批量处理Excel文件")
        print("  --type [ckin|tn|rtu1|all]: 指定查询类型（默认: ckin）")
        print("\n示例:")
        print("  python pnr_processor_new.py --pnr HZMZ28 --get-tn")
        print("  python pnr_processor_new.py --pnr HZMZ28 --get-rtu1")
        print("  python pnr_processor_new.py --pnr KM26TL --get-rtu1 --extract 1612ZX")
        print("  python pnr_processor_new.py --ticket 7849534336809 --get-tn")
        print("  python pnr_processor_new.py --input input.xlsx --output output.xlsx --type tn")
        print("  python pnr_processor_new.py --input input.xlsx --output output.xlsx --type rtu1")


if __name__ == '__main__':
    main()