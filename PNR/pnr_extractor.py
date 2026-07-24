import requests
import json
import re
from typing import Optional, Dict
import openpyxl
from openpyxl import load_workbook

# 配置
API_URL = 'https://eterm1w-api.etriplink.com/ibe-proxy/ota/xml/AirResRetCompleteInfo'
HEADERS = {
    'Aid': '84403',
    'Content-Type': 'application/json',
    'Authorization': 'Basic a21nMzE5enN5OnludGIxMjU4MA==',
}


def query_pnr(pnr: str, office: str = "KMG319") -> Optional[Dict]:
    """
    查询PNR信息

    Args:
        pnr: PNR号
        office: 办公点代码，默认为KMG319

    Returns:
        响应JSON数据，失败返回None
    """
    payload = {
        "RetrieveCRSPNRInfoHistoryRQ": {
            "Request": {
                "office": office,
                "pnr": pnr
            }
        }
    }

    try:
        response = requests.post(API_URL, headers=HEADERS, json=payload, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        print(f"查询PNR {pnr} 失败: {str(e)}")
        return None


def extract_gp_code(endorsement_infos: list) -> Optional[str]:
    """
    从endorsement_infos中提取*GP开头的值

    Args:
        endorsement_infos: endorsementInfos数组

    Returns:
        提取到的GP代码，如: "*GP1612ZXKM"，未找到返回None
    """
    if not endorsement_infos:
        return None

    for item in endorsement_infos:
        text = item.get('text', '')
        if text:
            # 匹配 *GP 开头的值，可能是完整字符串或包含*GP的字符串
            # 优先提取完整的*GP字符串，支持 *GP 或 **GP 格式
            gp_match = re.search(r'\*{1,2}GP[A-Z0-9]+', text)
            if gp_match:
                return gp_match.group(0).strip()

    return None


def extract_vico_values(ssrs: list) -> Dict[str, Optional[str]]:
    """
    从ssrs中提取VICO开头的A值、B值和D值

    Args:
        ssrs: ssrs数组

    Returns:
        包含A值、B值、D值的字典: {'vico_a': str, 'vico_b': str, 'vico_d': str}
        未找到的值返回None

    VICO值分类规则:
        - A值: VICOGP 或 VICOGP+数字（后缀以GP开头）
        - B值: VICO+两位数字+GP+数字（后缀不以GP开头，中间含GP）
        - D值: VICO开头但不包含GP（后缀不以GP开头且不含GP）
    """
    result = {
        'vico_a': None,
        'vico_b': None,
        'vico_d': None
    }

    if not ssrs:
        return result

    for item in ssrs:
        # 只处理CKIN类型的SSR
        ssr_code = item.get('ssrCode', '')
        if ssr_code != 'CKIN':
            continue

        text = item.get('text', '')
        if not text:
            continue

        # 匹配所有 VICO 开头的值
        all_vico = re.findall(r'VICO[A-Z0-9]+', text)
        if not all_vico:
            continue

        for match in all_vico:
            # A值: VICOGP 或 VICOGP+数字（后缀以GP开头）
            if re.match(r'^VICOGP\d*$', match):
                result['vico_a'] = match
            # B值: VICO+两位数字+GP+数字（后缀不以GP开头，中间含GP）
            elif re.match(r'^VICO\d{2}GP\d+$', match):
                result['vico_b'] = match
            # D值: VICO开头但不包含GP（后缀不以GP开头且不含GP）
            elif 'GP' not in match:
                result['vico_d'] = match

    return result


def extract_fp_c_value(text: str) -> Optional[str]:
    """
    从文本中提取FP开头的C值

    Args:
        text: 包含FP信息的文本

    Returns:
        提取到的FP开头的值，如: "FP/CASH,CNY/*KMG1234567"
        匹配格式: FP/CASH,CNY/*KMG******* 或 *KMG*******（有七位数字）
        未找到返回None
    """
    if not text:
        return None

    # 模式1: FP/CASH,CNY/*KMGxxxxxxx (完整格式)
    fp_match = re.search(r'FP/CASH,CNY/(\*KMG\d{7})', text)
    if fp_match:
        # 返回完整的FP/CASH,CNY/*KMGxxxxxxx格式
        return f"FP/CASH,CNY/{fp_match.group(1)}"

    # 如果完整格式没找到，尝试匹配简化格式 *KMGxxxxxxx
    kmg_match = re.search(r'(\*KMG\d{7})', text)
    if kmg_match:
        return kmg_match.group(1)

    # 还可以尝试更宽松的FP开头匹配
    fp_loose_match = re.search(r'FP[^/\s]*[/\s]*[^/\s]*[/\s]*(\*KMG\d{7})', text)
    if fp_loose_match:
        return fp_loose_match.group(1)

    return None


def extract_vico_e(text: str) -> Optional[str]:
    """
    从 FP/CASH,CNY/* 文本中提取 E 值

    E 值数据格式例如：
        FP/CASH,CNY/*GN0135KM     → E = GN0135KM
        FP/CASH,CNY/*KMG1234567   → 这是 C 值，不是 E 值
        两者可能同时出现在原文中

    Args:
        text: 原始 originalRTC 文本

    Returns:
        提取到的 E 值，未找到返回 None
    """
    if not text:
        return None

    # 模式1: FP/CASH,CNY/*KMG\d{7}/ 后面的值（C值在前 / E值在后）
    m1 = re.search(r'FP/CASH,CNY/\*KMG\d{7}/([A-Z0-9]+)', text)
    if m1:
        return m1.group(1)

    # 模式2: FP/CASH,CNY/* 后面直接跟非 KMG 的值（即 E 值本身）
    # 例如: FP/CASH,CNY/*GN0135KM → 提取 GN0135KM
    m2 = re.search(r'FP/CASH,CNY/\*((?!KMG\d)[A-Z0-9]+)', text)
    if m2:
        return m2.group(1)

    return None


def process_pnr_response(response_data: Dict) -> Dict:
    """
    处理PNR响应数据，提取关键信息

    Args:
        response_data: API返回的JSON数据

    Returns:
        包含提取信息的字典
    """
    result = {
        'pnr': None,
        'gp_code': None,
        'vico_a': None,
        'vico_b': None,
        'vico_d': None,
        'fp_c_value': None,
        'vico_e': None,
        'order_id': None,
        'status': 'error'
    }

    try:
        # 提取基础信息
        response = response_data.get('RetrieveCRSPNRInfoHistroryRS', {})
        result_obj = response.get('Response', {}).get('result', {})
        order_info = result_obj.get('orderInfoC', {})

        if result_obj.get('retrieveStatus') != 'SUCCESS':
            result['status'] = 'failed'
            return result

        result['pnr'] = order_info.get('crsPnrLocator')
        result['order_id'] = order_info.get('orderId')

        # 获取PNR对象
        pnr_data = order_info.get('pnr', {})

        # 提取*GP代码 - 从多个位置查找
        gp_code_found = False

        # 方法1: 从histories中的endorsementInfos提取
        histories = pnr_data.get('histories', [])
        for history in histories:
            endorsement_infos = history.get('endorsementInfos', [])
            gp_code = extract_gp_code(endorsement_infos)
            if gp_code:
                result['gp_code'] = gp_code
                gp_code_found = True
                break

        # 方法2: 如果方法1没找到，从fareCalculations中查找
        if not gp_code_found:
            for history in histories:
                fare_calcs = history.get('fareCalculations', [])
                for fare_calc in fare_calcs:
                    text = fare_calc.get('text', '')
                    if text:
                        # 在fareCalculations中查找GP代码
                        # 可能的格式: **GP, *GPxxx, 或者 GPxxx
                        gp_match = re.search(r'\*{1,2}GP[A-Z0-9]*', text)
                        if gp_match:
                            gp_code = gp_match.group(0).strip()
                            # 如果是 **GP，需要从前面的文本中提取 GP 后面的内容
                            if gp_code == '**GP':
                                # 尝试从 RGPL 或类似模式中提取
                                rgp_match = re.search(r'RGPL?(\d+[A-Z]*)', text)
                                if rgp_match:
                                    result['gp_code'] = '*GP' + rgp_match.group(1)
                                    gp_code_found = True
                                    break
                            else:
                                result['gp_code'] = gp_code
                                gp_code_found = True
                                break
                if gp_code_found:
                    break

        # 提取FP开头的C值 - 从originalRTC中查找
        original_rtc = order_info.get('originalRTC', '')
        if original_rtc:
            fp_c_value = extract_fp_c_value(original_rtc)
            if fp_c_value:
                result['fp_c_value'] = fp_c_value
            vico_e = extract_vico_e(original_rtc)
            if vico_e:
                result['vico_e'] = vico_e

        # 提取VICO值
        ssrs = pnr_data.get('ssrs', [])
        vico_values = extract_vico_values(ssrs)
        result['vico_a'] = vico_values['vico_a']
        result['vico_b'] = vico_values['vico_b']
        result['vico_d'] = vico_values['vico_d']

        result['status'] = 'success'

    except Exception as e:
        print(f"处理响应数据时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        result['status'] = 'error'

    return result


def query_and_extract(pnr: str, office: str = "KMG319", debug: bool = False) -> Dict:
    """
    查询PNR并提取关键信息

    Args:
        pnr: PNR号
        office: 办公点代码，默认为KMG319
        debug: 是否输出调试信息

    Returns:
        包含提取信息的字典
    """
    print(f"正在查询PNR: {pnr}")
    response = query_pnr(pnr, office)
    if response:
        if debug:
            print("\n完整响应数据:")
            print(json.dumps(response, ensure_ascii=False, indent=2))
        result = process_pnr_response(response)
        print(f"查询完成: PNR={result['pnr']}, GP代码={result['gp_code']}, "
              f"VICO_A={result['vico_a']}, VICO_B={result['vico_b']}, VICO_D={result['vico_d']}, VICO_E={result['vico_e']}, FP_C值={result['fp_c_value']}")
        return result
    else:
        return {
            'pnr': pnr,
            'gp_code': None,
            'vico_a': None,
            'vico_b': None,
            'vico_d': None,
            'vico_e': None,
            'fp_c_value': None,
            'order_id': None,
            'status': 'query_failed'
        }


def format_result(result: Dict) -> str:
    """
    格式化输出结果

    Args:
        result: 查询结果字典

    Returns:
        格式化字符串
    """
    return (f"PNR: {result['pnr']}\n"
            f"订单ID: {result['order_id']}\n"
            f"GP代码: {result['gp_code'] or '未找到'}\n"
            f"VICO_A值: {result['vico_a'] or '未找到'}\n"
            f"VICO_B值: {result['vico_b'] or '未找到'}\n"
            f"VICO_D值: {result['vico_d'] or '未找到'}\n"
            f"VICO_E值: {result['vico_e'] or '未找到'}\n"
            f"FP_C值: {result['fp_c_value'] or '未找到'}\n"
            f"状态: {result['status']}")


def process_excel_file(input_file: str, output_file: str, office: str = "KMG319") -> bool:
    """
    批量处理Excel文件中的PNR

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径
        office: 办公点代码，默认为KMG319

    Returns:
        处理成功返回True，失败返回False
    """
    try:
        # 读取Excel文件
        wb = load_workbook(input_file)
        ws = wb.active

        # 添加表头（如果还没有的话）
        headers = ['PNR', '订单ID', 'GP代码', 'VICO_A值', 'VICO_B值', 'VICO_D值', 'VICO_E值', 'FP_C值', '状态']
        if ws[1][0].value != 'PNR':
            ws.insert_rows(1)
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        else:
            # 确保有足够的列
            for col, header in enumerate(headers, 1):
                if ws.cell(row=1, column=col).value != header:
                    ws.cell(row=1, column=col, value=header)

        # 统计
        total_rows = ws.max_row - 1  # 减去表头
        processed = 0
        success_count = 0
        failed_count = 0

        print(f"\n开始处理Excel文件: {input_file}")
        print(f"共有 {total_rows} 个PNR需要处理")
        print("=" * 60)

        # 处理每一行
        for row_idx in range(2, ws.max_row + 1):
            pnr = ws.cell(row=row_idx, column=1).value

            if not pnr:
                print(f"第 {row_idx} 行: PNR为空，跳过")
                continue

            pnr = str(pnr).strip().upper()  # 转大写

            print(f"\n[{processed + 1}/{total_rows}] 处理PNR: {pnr}")

            # 查询并提取
            result = query_and_extract(pnr, office)

            # 写入结果
            ws.cell(row=row_idx, column=2, value=result['order_id'])
            ws.cell(row=row_idx, column=3, value=result['gp_code'])
            ws.cell(row=row_idx, column=4, value=result['vico_a'])
            ws.cell(row=row_idx, column=5, value=result['vico_b'])
            ws.cell(row=row_idx, column=6, value=result['vico_d'])
            ws.cell(row=row_idx, column=7, value=result['vico_e'])
            ws.cell(row=row_idx, column=8, value=result['fp_c_value'])
            ws.cell(row=row_idx, column=9, value=result['status'])

            processed += 1

            if result['status'] == 'success':
                success_count += 1
                print(f"  [成功] GP代码={result['gp_code']}, VICO_A={result['vico_a']}, VICO_B={result['vico_b']}, VICO_D={result['vico_d']}, VICO_E={result['vico_e']}, FP_C值={result['fp_c_value']}")
            else:
                failed_count += 1
                print(f"  [失败] {result['status']}")

        # 保存结果
        wb.save(output_file)

        # 输出统计信息
        print("\n" + "=" * 60)
        print(f"处理完成！")
        print(f"总计: {processed} 个PNR")
        print(f"成功: {success_count} 个")
        print(f"失败: {failed_count} 个")
        print(f"结果已保存到: {output_file}")
        print("=" * 60)

        return True

    except Exception as e:
        print(f"\n处理Excel文件时出错: {str(e)}")
        return False


def create_template_excel(output_file: str = "pnr_template.xlsx"):
    """
    创建PNR批量导入模板Excel文件

    Args:
        output_file: 输出文件路径
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PNR批量导入"

        # 设置表头
        headers = ['PNR', '订单ID', 'GP代码', 'VICO_A值', 'VICO_B值', 'VICO_D值', 'VICO_E值', 'FP_C值', '状态']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 添加示例数据
        example_data = ['KF1XHV', '', '', '', '', '', '', '', '']
        for col, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=value)

        # 保存文件
        wb.save(output_file)
        print(f"模板文件已创建: {output_file}")
        print("\n使用说明:")
        print("1. 打开Excel文件")
        print("2. 在'PNR'列填入需要查询的PNR号")
        print("3. 保存文件")
        print("4. 使用以下命令批量处理:")
        print(f"   python pnr_extractor.py --input {output_file} --output result.xlsx")

    except Exception as e:
        print(f"创建模板文件时出错: {str(e)}")


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='从PNR响应中提取GP代码和VICO值')
    parser.add_argument('pnr', nargs='?', help='PNR号（单个查询时使用）')
    parser.add_argument('--office', default='KMG319', help='办公点代码（默认: KMG319）')
    parser.add_argument('--input', help='输入Excel文件路径（批量查询时使用）')
    parser.add_argument('--output', help='输出Excel文件路径（批量查询时使用）')
    parser.add_argument('--create-template', action='store_true', help='创建PNR批量导入模板')
    parser.add_argument('--debug', action='store_true', help='输出完整的响应数据用于调试')

    args = parser.parse_args()

    # 创建模板
    if args.create_template:
        create_template_excel()

    # 批量处理Excel
    elif args.input and args.output:
        process_excel_file(args.input, args.output, args.office)

    # 单个PNR查询
    elif args.pnr:
        result = query_and_extract(args.pnr, args.office, debug=args.debug)

        # 输出结果
        print("\n" + "=" * 60)
        print(format_result(result))
        print("=" * 60)

        # 输出JSON格式
        print("\nJSON格式:")
        print(json.dumps(result, ensure_ascii=False, indent=2))

    else:
        # 显示使用帮助
        print("使用说明:")
        print("\n1. 单个PNR查询:")
        print("   python pnr_extractor.py KF1XHV")
        print("   python pnr_extractor.py JYPY5Z --office KMG319")
        print("   python pnr_extractor.py JYPY5Z --debug  # 查看完整响应数据")
        print("\n2. 批量处理Excel:")
        print("   python pnr_extractor.py --input input.xlsx --output result.xlsx")
        print("\n3. 创建PNR导入模板:")
        print("   python pnr_extractor.py --create-template")
