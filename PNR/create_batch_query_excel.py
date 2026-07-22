# -*- coding: utf-8 -*-
"""
PNR批量查询Excel模板创建脚本
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "PNR批量查询"

# 定义样式
header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

input_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色背景表示需要填写

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 设置表头
headers = ['PNR', '订单ID', 'GP代码', 'VICO_A值', 'VICO_B值', 'FP_C值', '状态']
column_widths = [15, 25, 20, 18, 18, 25, 15]

for col, (header, width) in enumerate(zip(headers, column_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws.column_dimensions[cell.column_letter].width = width

# 设置行高
ws.row_dimensions[1].height = 25

# 添加示例PNR（可选）
example_pnrs = ['KF1XHV', 'KRVCDG', 'JQBYX9']
for row, pnr in enumerate(example_pnrs, 2):
    ws.cell(row=row, column=1, value=pnr)
    ws.cell(row=row, column=1).fill = input_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

# 为其他列添加边框
for row in range(2, 5):
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = thin_border

# 添加使用说明
ws['A8'] = '使用说明:'
ws['A8'].font = Font(bold=True, size=11)
instructions = [
    '1. 在"PNR"列填写需要查询的PNR号（黄色单元格）',
    '2. 保存文件',
    '3. 运行批量查询: python batch_query.py',
    '4. 查询结果将自动填充到对应列'
]
for i, instruction in enumerate(instructions, 9):
    ws[f'A{i}'] = instruction
    ws[f'A{i}'].font = Font(size=10)

# 保存文件
output_path = r'c:\Users\Administrator\OneDrive\桌面\MU,8L大客户打标脚本\PNR\pnr_batch_query.xlsx'
wb.save(output_path)

print(f"PNR批量查询模板已创建: {output_path}")
print("\n模板包含:")
print("  - PNR列: 填写需要查询的PNR号（黄色背景）")
print("  - 订单ID: 查询结果自动填充")
print("  - GP代码: 查询结果自动填充")
print("  - VICO_A值: 查询结果自动填充")
print("  - VICO_B值: 查询结果自动填充")
print("  - FP_C值: 查询结果自动填充")
print("  - 状态: 查询状态（success/failed/error）")
