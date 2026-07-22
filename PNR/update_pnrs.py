import openpyxl

# 读取模板
wb = openpyxl.load_workbook('pnr_template.xlsx')
ws = wb.active

# 添加有效的PNR
pnrs_to_test = ['KRVCDG', 'JQBYX9', 'KF1XHV']

# 清空数据行(保留表头)
for row in range(2, ws.max_row + 1):
    for col in range(1, 8):
        ws.cell(row=row, column=col, value=None)

# 填入PNR
for idx, pnr in enumerate(pnrs_to_test, start=2):
    ws.cell(row=idx, column=1, value=pnr)
    print(f"添加PNR: {pnr}")

# 保存
wb.save('pnr_template.xlsx')
print(f"\n已保存3个PNR到pnr_template.xlsx")
