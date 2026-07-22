import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('c:/Users/Administrator/CodeBuddy/20260324101611/pnr_template.xlsx')
    print('文件读取成功')
    print(f'工作表数量: {len(wb.sheetnames)}')
    ws = wb.active
    print(f'工作表名称: {ws.title}')
    print(f'最大行数: {ws.max_row}')
    print(f'最大列数: {ws.max_column}')
except Exception as e:
    print(f'错误: {e}')
    sys.exit(1)
