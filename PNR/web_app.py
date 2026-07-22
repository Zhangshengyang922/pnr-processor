# -*- coding: utf-8 -*-
"""
PNR查询 Web 应用
基于 Flask 的 PNR 数据提取工具 Web 界面
"""

import os
import sys
import io
import json
import re
import time
from datetime import datetime
from typing import Optional, Dict

import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, render_template, request, jsonify, send_file, session

# ==================== PyInstaller 路径处理 ====================
def resource_path(relative_path):
    """获取资源文件的绝对路径，兼容 PyInstaller 打包"""
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

# ==================== 配置 ====================
API_URL = 'https://eterm1w-api.etriplink.com/ibe-proxy/ota/xml/AirResRetCompleteInfo'
HEADERS = {
    'Aid': '84403',
    'Content-Type': 'application/json',
    'Authorization': 'Basic a21nMzE5enN5OnludGIxMjU4MA==',
}

app = Flask(__name__, template_folder=resource_path('templates'))
app.secret_key = 'pnr_processor_secret_key_2024'


# ==================== PNR 核心逻辑（来自 pnr_extractor.py） ====================

def query_pnr(pnr: str, office: str = "KMG319") -> tuple[Optional[Dict], Optional[str]]:
    """查询PNR，返回 (响应数据, 错误信息)"""
    payload = {
        "RetrieveCRSPNRInfoHistoryRQ": {
            "Request": {"office": office, "pnr": pnr}
        }
    }
    try:
        response = requests.post(API_URL, headers=HEADERS, json=payload, timeout=15)
        response.raise_for_status()
        return response.json(), None
    except requests.HTTPError as e:
        return None, f"HTTP {response.status_code}: {response.text[:200]}"
    except requests.Timeout:
        return None, "请求超时（15秒）"
    except requests.ConnectionError:
        return None, "网络连接失败，请检查网络"
    except requests.RequestException as e:
        return None, f"请求异常: {str(e)[:200]}"


def extract_gp_code(endorsement_infos: list) -> Optional[str]:
    if not endorsement_infos:
        return None
    for item in endorsement_infos:
        text = item.get('text', '')
        if text:
            gp_match = re.search(r'\*{1,2}GP[A-Z0-9]+', text)
            if gp_match:
                return gp_match.group(0).strip()
    return None


def extract_vico_values(ssrs: list) -> Dict[str, Optional[str]]:
    result = {'vico_a': None, 'vico_b': None}
    if not ssrs:
        return result
    for item in ssrs:
        if item.get('ssrCode', '') != 'CKIN':
            continue
        text = item.get('text', '')
        if not text:
            continue
        vico_matches = re.findall(r'VICO\d*GP\d*', text)
        for match in vico_matches:
            if match == 'VICOGP' or re.match(r'^VICOGP\d+$', match):
                result['vico_a'] = match
            elif re.match(r'^VICO\d{2}GP\d+$', match):
                result['vico_b'] = match
    return result


def extract_fp_c_value(text: str) -> Optional[str]:
    if not text:
        return None
    fp_match = re.search(r'FP/CASH,CNY/(\*KMG\d{7})', text)
    if fp_match:
        return f"FP/CASH,CNY/{fp_match.group(1)}"
    kmg_match = re.search(r'(\*KMG\d{7})', text)
    if kmg_match:
        return kmg_match.group(1)
    return None


def process_pnr_response(response_data: Dict) -> Dict:
    result = {
        'pnr': None, 'gp_code': None, 'vico_a': None,
        'vico_b': None, 'fp_c_value': None, 'order_id': None,
        'status': 'error', 'error_msg': None
    }
    try:
        response = response_data.get('RetrieveCRSPNRInfoHistroryRS', {})
        result_obj = response.get('Response', {}).get('result', {})
        order_info = result_obj.get('orderInfoC', {})

        # 优先从响应中提取PNR号
        result['pnr'] = order_info.get('crsPnrLocator')

        if result_obj.get('retrieveStatus') != 'SUCCESS':
            result['status'] = 'failed'
            # 提取API错误信息
            errors = response.get('Errors', [])
            if errors:
                result['error_msg'] = errors[0].get('ShortText', 'API返回失败状态')
            else:
                result['error_msg'] = result_obj.get('message') or result_obj.get('retrieveStatus') or 'API查询失败'
            return result

        result['order_id'] = order_info.get('orderId')
        pnr_data = order_info.get('pnr', {})

        # 提取 GP 代码
        gp_code_found = False
        histories = pnr_data.get('histories', [])
        for history in histories:
            endorsement_infos = history.get('endorsementInfos', [])
            gp_code = extract_gp_code(endorsement_infos)
            if gp_code:
                result['gp_code'] = gp_code
                gp_code_found = True
                break

        if not gp_code_found:
            for history in histories:
                fare_calcs = history.get('fareCalculations', [])
                for fare_calc in fare_calcs:
                    text = fare_calc.get('text', '')
                    if text:
                        gp_match = re.search(r'\*{1,2}GP[A-Z0-9]*', text)
                        if gp_match:
                            gp_code = gp_match.group(0).strip()
                            if gp_code == '**GP':
                                rgp_match = re.search(r'RGPL?(\d+[A-Z]*)', text)
                                if rgp_match:
                                    result['gp_code'] = '*GP' + rgp_match.group(1)
                                    gp_code_found = True
                                    break
                            else:
                                result['gp_code'] = gp_code
                                gp_code_found = True
                                break
                if gp_code_found:
                    break

        # 提取 FP/C 值
        original_rtc = order_info.get('originalRTC', '')
        if original_rtc:
            fp_c_value = extract_fp_c_value(original_rtc)
            if fp_c_value:
                result['fp_c_value'] = fp_c_value

        # 提取 VICO 值
        ssrs = pnr_data.get('ssrs', [])
        vico_values = extract_vico_values(ssrs)
        result['vico_a'] = vico_values['vico_a']
        result['vico_b'] = vico_values['vico_b']

        result['status'] = 'success'
    except Exception as e:
        result['status'] = 'error'
        result['error_msg'] = f'解析异常: {str(e)[:100]}'
    return result


def query_and_extract(pnr: str, office: str = "KMG319") -> Dict:
    response, error_msg = query_pnr(pnr, office)
    if response:
        result = process_pnr_response(response)
        # 如果API返回了PNR号就用，否则用查询时传入的
        if not result.get('pnr'):
            result['pnr'] = pnr
        return result
    else:
        return {
            'pnr': pnr, 'gp_code': None, 'vico_a': None,
            'vico_b': None, 'fp_c_value': None, 'order_id': None,
            'status': 'query_failed', 'error_msg': error_msg
        }


# ==================== Web 页面路由 ====================

@app.route('/')
def index():
    """主页面"""
    return render_template('index.html')


# ==================== API 路由 ====================

@app.route('/api/query_single', methods=['POST'])
def api_query_single():
    """单个 PNR 查询"""
    data = request.get_json()
    pnr = data.get('pnr', '').strip().upper()
    office = data.get('office', 'KMG319').strip()

    if not pnr:
        return jsonify({'success': False, 'message': '请输入PNR号'})

    if not re.match(r'^[A-Z0-9]{5,6}$', pnr):
        return jsonify({'success': False, 'message': 'PNR号格式不正确，应为5-6位字母或数字'})

    result = query_and_extract(pnr, office)
    return jsonify({'success': True, 'data': result})


@app.route('/api/query_batch', methods=['POST'])
def api_query_batch():
    """批量 PNR 查询（上传 Excel）"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请上传Excel文件'})

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传.xlsx或.xls格式的文件'})

    office = request.form.get('office', 'KMG319').strip()

    try:
        # 读取上传的 Excel
        wb = load_workbook(file)
        ws = wb.active

        # 读取第一列的 PNR 号
        pnr_list = []
        for row_idx in range(2, ws.max_row + 1):
            pnr = ws.cell(row=row_idx, column=1).value
            if pnr:
                pnr = str(pnr).strip().upper()
                if re.match(r'^[A-Z0-9]{5,6}$', pnr):
                    pnr_list.append(pnr)

        if not pnr_list:
            return jsonify({'success': False, 'message': '未在Excel第一列中找到有效的PNR号'})

        # 创建输出 Excel
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "PNR查询结果"

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        headers = ['PNR', '订单ID', 'GP代码', 'VICO_A值', 'VICO_B值', 'FP_C值', '状态']
        col_widths = [15, 25, 20, 18, 18, 28, 14]
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = out_ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            out_ws.column_dimensions[cell.column_letter].width = w

        # 逐行查询
        results = []
        success_count = 0
        for idx, pnr in enumerate(pnr_list):
            result = query_and_extract(pnr, office)
            results.append(result)

            row = idx + 2
            values = [result['pnr'], result['order_id'], result['gp_code'],
                       result['vico_a'], result['vico_b'], result['fp_c_value'], result['status']]
            for col, val in enumerate(values, 1):
                cell = out_ws.cell(row=row, column=col, value=val if val else '')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 状态着色
            status_cell = out_ws.cell(row=row, column=7)
            if result['status'] == 'success':
                status_cell.fill = success_fill
                success_count += 1
            else:
                status_cell.fill = fail_fill

        # 保存到内存
        output = io.BytesIO()
        out_wb.save(output)
        output.seek(0)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'PNR查询结果_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'success': False, 'message': f'处理文件时出错: {str(e)}'})


@app.route('/api/progress_query_batch', methods=['POST'])
def api_progress_query_batch():
    """批量查询（带进度返回 - 用于前端实时显示）"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请上传Excel文件'})

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传.xlsx或.xls格式的文件'})

    office = request.form.get('office', 'KMG319').strip()

    try:
        wb = load_workbook(file)
        ws = wb.active
        pnr_list = []
        for row_idx in range(2, ws.max_row + 1):
            pnr = ws.cell(row=row_idx, column=1).value
            if pnr is not None:
                pnr = str(pnr).strip().upper()
                # 去掉可能的空格和换行
                pnr = re.sub(r'\s+', '', pnr)
                if re.match(r'^[A-Z0-9]{5,6}$', pnr):
                    pnr_list.append(pnr)

        if not pnr_list:
            return jsonify({'success': False, 'message': f'未找到有效PNR号（共扫描{ws.max_row-1}行，第一列无5-6位字母数字组合）'})

        results = []
        for pnr in pnr_list:
            result = query_and_extract(pnr, office)
            results.append(result)

        # 收集错误摘要
        error_summary = {}
        for r in results:
            if r.get('error_msg'):
                msg = r['error_msg']
                error_summary[msg] = error_summary.get(msg, 0) + 1

        return jsonify({
            'success': True,
            'total': len(results),
            'success_count': sum(1 for r in results if r['status'] == 'success'),
            'data': results,
            'error_summary': error_summary
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'处理出错: {str(e)}'})


@app.route('/api/create_template')
def api_create_template():
    """下载 PNR 导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PNR批量查询"

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    input_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['PNR', '订单ID', 'GP代码', 'VICO_A值', 'VICO_B值', 'FP_C值', '状态']
    col_widths = [15, 25, 20, 18, 18, 28, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    examples = ['KF1XHV', 'KRVCDG', 'JQBYX9']
    for r, pnr in enumerate(examples, 2):
        cell = ws.cell(row=r, column=1, value=pnr)
        cell.fill = input_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        for c in range(2, 8):
            ws.cell(row=r, column=c).border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='PNR批量查询模板.xlsx'
    )


# ==================== 启动 ====================

if __name__ == '__main__':
    # 开发模式下确保 templates 目录存在
    if not getattr(sys, 'frozen', False):
        os.makedirs(resource_path('templates'), exist_ok=True)

    import webbrowser
    import threading

    port = 5000
    url = f"http://127.0.0.1:{port}"

    print("=" * 60)
    print("  PNR数据提取工具 - Web版")
    print(f"  浏览器访问: {url}")
    print("=" * 60)

    # 延迟1秒后自动打开浏览器
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()

    app.run(debug=False, host='0.0.0.0', port=port, threaded=True)
