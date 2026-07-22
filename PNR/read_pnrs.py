import openpyxl

# 读取CodeBuddy目录下的模板文件
wb = openpyxl.load_workbook('c:/Users/Administrator/CodeBuddy/20260324101611/pnr_template.xlsx')
ws = wb.active

print("读取到的PNR列表:")
print("=" * 60)

for row_idx in range(2, ws.max_row + 1):
    pnr = ws.cell(row=row_idx, column=1).value
    if pnr:
        print(f"  第{row_idx}行: {pnr}")

print("=" * 60)
print(f"共 {ws.max_row - 1} 个PNR")
